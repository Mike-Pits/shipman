import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime
import os

class ExcelExporter:
    """Excel export without pandas dependency"""
    
    @staticmethod
    def export_report(data, filename, title, headers):
        """Export report data to Excel"""
        os.makedirs('exports', exist_ok=True)
        
        filepath = os.path.join('exports', filename)
        if not filepath.endswith('.xlsx'):
            filepath += '.xlsx'
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Report"
        
        # Title
        ws.merge_cells(f'A1:{get_column_letter(len(headers))}1')
        title_cell = ws['A1']
        title_cell.value = title
        title_cell.font = Font(size=14, bold=True)
        title_cell.alignment = Alignment(horizontal='center')
        
        # Date
        ws.merge_cells(f'A2:{get_column_letter(len(headers))}2')
        date_cell = ws['A2']
        date_cell.value = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        date_cell.font = Font(size=10, italic=True)
        date_cell.alignment = Alignment(horizontal='center')
        
        # Headers
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col_idx)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
        
        # Data
        for row_idx, row_data in enumerate(data, 5):
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        # Auto-size columns
        for col_idx in range(1, len(headers) + 1):
            column_letter = get_column_letter(col_idx)
            max_length = len(headers[col_idx - 1])
            for row in range(5, len(data) + 5):
                cell_value = ws.cell(row=row, column=col_idx).value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
        
        wb.save(filepath)
        return filepath

excel_exporter = ExcelExporter()